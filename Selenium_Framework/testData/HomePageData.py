import openpyxl


class HomePageData:
    test_HomePage_data = [{"FirstName": "Shank", "Email": "Karan@gmail.com", "Gender": "Male"},
                          {"FirstName": "Cherry", "Email": "Charneet@gmail.com", "Gender": "Female"}]

    @staticmethod
    def getTestData(test_Case_Name):
        Dict = {}
        book = openpyxl.load_workbook(
            "/home/shanky/PycharmProjects/Python_Testing_Selenium/Selenium_Framework/PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # Rows
            if sheet.cell(row=i, column=1).value == test_Case_Name:
                for j in range(2, sheet.max_column + 1):  # Columns
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]
