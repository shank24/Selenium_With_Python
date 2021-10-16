import openpyxl

book = openpyxl.load_workbook("/home/shanky/PycharmProjects/Python_Testing_Selenium/Selenium_Framework/PythonDemo.xlsx")
sheet = book.active
#initializing Empty Dictionary
Dict = {}


cell = sheet.cell(row=1 , column=2)
print(cell.value)
sheet.cell(row=2 , column=2).value = "Shank"

print(sheet.cell(row=2 , column=2).value)

print(sheet.max_row)
print(sheet.max_column)

print (sheet['A5'].value)

for i in range(1, sheet.max_row+1):  #Rows
    if sheet.cell(row=i, column=1).value == "Test2":
        for j in range(2, sheet.max_column+1):  #Columns
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value


        print(Dict)